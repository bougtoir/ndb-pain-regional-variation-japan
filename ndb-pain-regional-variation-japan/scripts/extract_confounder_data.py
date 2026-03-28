#!/usr/bin/env python3
"""Extract confounder disease proxy data from NDB open data for CPSP analysis."""
import openpyxl
import csv
import os
import re

OUTPUT_DIR = '/home/ubuntu/analysis/output/'
DATA_DIR = '/home/ubuntu/analysis/data/'

# Prefecture names mapping (code -> name)
PREF_NAMES = {
    1:'北海道',2:'青森県',3:'岩手県',4:'宮城県',5:'秋田県',6:'山形県',7:'福島県',
    8:'茨城県',9:'栃木県',10:'群馬県',11:'埼玉県',12:'千葉県',13:'東京都',14:'神奈川県',
    15:'新潟県',16:'富山県',17:'石川県',18:'福井県',19:'山梨県',20:'長野県',
    21:'岐阜県',22:'静岡県',23:'愛知県',24:'三重県',25:'滋賀県',26:'京都府',
    27:'大阪府',28:'兵庫県',29:'奈良県',30:'和歌山県',31:'鳥取県',32:'島根県',
    33:'岡山県',34:'広島県',35:'山口県',36:'徳島県',37:'香川県',38:'愛媛県',
    39:'高知県',40:'福岡県',41:'佐賀県',42:'長崎県',43:'熊本県',44:'大分県',
    45:'宮崎県',46:'鹿児島県',47:'沖縄県'
}

# Column mapping: columns 9-55 = prefectures 01-47 (0-indexed: col 9 = 北海道, col 10 = 青森...)
# Based on header row: col 8 = 総計, col 9 = 01北海道, col 10 = 02青森県, ...
PREF_COL_START = 9  # 0-indexed column for 北海道

def safe_float(val):
    """Convert value to float, handling '-' and None."""
    if val is None or val == '-' or val == '－':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def extract_drug_data_by_prefecture(filepath, sheet_name, drug_keywords, label):
    """Extract drug prescription quantities by prefecture from NDB outpatient drug data."""
    print(f"\n--- Extracting {label} from {sheet_name} ---")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb[sheet_name]
    
    # Initialize prefecture totals
    pref_totals = {i: 0.0 for i in range(1, 48)}
    total_national = 0.0
    matched_drugs = []
    
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        if len(vals) < 56:
            continue
        drug_name = str(vals[3]) if vals[3] else ''
        if any(kw in drug_name for kw in drug_keywords):
            matched_drugs.append(drug_name)
            national = safe_float(vals[8])
            total_national += national
            for pref_code in range(1, 48):
                col_idx = PREF_COL_START + (pref_code - 1)
                if col_idx < len(vals):
                    pref_totals[pref_code] += safe_float(vals[col_idx])
    
    wb.close()
    print(f"  Matched {len(matched_drugs)} drug entries")
    print(f"  National total: {total_national:,.0f}")
    if matched_drugs[:5]:
        print(f"  Examples: {matched_drugs[:5]}")
    return pref_totals, total_national

def extract_procedure_data_by_prefecture(filepath, sheet_name, proc_keywords, label):
    """Extract procedure counts by prefecture from NDB procedure data."""
    print(f"\n--- Extracting {label} from {sheet_name} ---")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb[sheet_name]
    
    pref_totals = {i: 0.0 for i in range(1, 48)}
    total_national = 0.0
    matched_procs = []
    
    # Need to understand the column structure of this file first
    header_row = None
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        if i < 5:
            print(f"  Header row {i}: {vals[:8]}")
        
        # Check if any column contains the procedure keyword
        proc_name = ''
        for ci in range(min(6, len(vals))):
            if vals[ci] and any(kw in str(vals[ci]) for kw in proc_keywords):
                proc_name = str(vals[ci])
                break
        
        if proc_name:
            matched_procs.append(proc_name)
            # Try to find prefecture data - need to understand column layout
            # For anesthesia/procedure files, the structure may differ
    
    wb.close()
    print(f"  Matched {len(matched_procs)} procedure entries")
    return pref_totals, total_national


# ============================================================
# STEP 1: Extract neuropathic pain drugs (CPSP proxy)
# ============================================================
print("=" * 80)
print("STEP 1: Neuropathic pain drugs (CPSP proxy)")
print("=" * 80)

outpatient_file = DATA_DIR + 'outpatient_drugs_prefecture.xlsx'

# 1a. Pregabalin family
pregabalin_kw = ['プレガバリン', 'リリカ']
pregabalin_data, pregabalin_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', pregabalin_kw, 'Pregabalin (all forms)')

# 1b. Mirogabalin family  
mirogabalin_kw = ['ミロガバリン', 'タリージェ']
mirogabalin_data, mirogabalin_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', mirogabalin_kw, 'Mirogabalin (all forms)')

# 1c. Duloxetine family (SNRI - also for pain)
duloxetine_kw = ['デュロキセチン', 'サインバルタ']
duloxetine_data, duloxetine_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', duloxetine_kw, 'Duloxetine (all forms)')

# 1d. Tramadol family
tramadol_kw = ['トラマドール', 'トラムセット']
tramadol_data, tramadol_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', tramadol_kw, 'Tramadol (all forms)')

# 1e. Neurotropin
neurotropin_kw = ['ノイロトロピン', 'ワクシニア']
neurotropin_data, neurotropin_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', neurotropin_kw, 'Neurotropin')

# Combine all neuropathic pain drugs
neuropathic_total = {}
for pc in range(1, 48):
    neuropathic_total[pc] = (pregabalin_data[pc] + mirogabalin_data[pc] + 
                              duloxetine_data[pc] + tramadol_data[pc] + neurotropin_data[pc])

# ============================================================
# STEP 2: Extract confounder disease proxies
# ============================================================
print("\n" + "=" * 80)
print("STEP 2: Confounder disease proxies")
print("=" * 80)

# 2a. Diabetes drugs (薬効分類 396: 糖尿病用剤)
# These include insulin, metformin, DPP-4 inhibitors, SGLT2 inhibitors, etc.
print("\n--- Searching for diabetes drugs (class 396) ---")
wb = openpyxl.load_workbook(outpatient_file, read_only=True)
sheet = wb['内服薬 外来 (院外)']
diabetes_data = {i: 0.0 for i in range(1, 48)}
diabetes_count = 0
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    vals = list(row)
    if len(vals) < 56:
        continue
    # Check drug class (column 0 or 1)
    drug_class = vals[0]
    drug_class_name = str(vals[1]) if vals[1] else ''
    if drug_class == 396 or '糖尿病' in drug_class_name:
        diabetes_count += 1
        for pref_code in range(1, 48):
            col_idx = PREF_COL_START + (pref_code - 1)
            if col_idx < len(vals):
                diabetes_data[pref_code] += safe_float(vals[col_idx])
wb.close()
print(f"  Found {diabetes_count} diabetes drug entries")
print(f"  National total: {sum(diabetes_data.values()):,.0f}")

# 2b. Herpes zoster antivirals (帯状疱疹治療薬)
# バラシクロビル, アシクロビル, ファムシクロビル, アメナメビル
herpes_kw = ['バラシクロビル', 'バルトレックス', 'アシクロビル', 'ゾビラックス',
             'ファムシクロビル', 'ファムビル', 'アメナメビル', 'アメナリーフ']
herpes_data, herpes_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', herpes_kw, 'Herpes zoster antivirals')

# 2c. Antidepressants (SSRI/SNRI/NaSSA/tricyclic - うつ病 proxy)
# Exclude duloxetine (already counted in neuropathic pain)
antidep_kw = ['パロキセチン', 'パキシル', 'セルトラリン', 'ジェイゾロフト',
              'エスシタロプラム', 'レクサプロ', 'フルボキサミン', 'デプロメール', 'ルボックス',
              'ミルナシプラン', 'トレドミン', 'ベンラファキシン', 'イフェクサー',
              'ミルタザピン', 'リフレックス', 'レメロン',
              'アミトリプチリン', 'トリプタノール', 'ノリトレン', 'イミプラミン', 'トフラニール',
              'クロミプラミン', 'アナフラニール', 'ボルチオキセチン', 'トリンテリックス']
antidep_data, antidep_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', antidep_kw, 'Antidepressants (excl. duloxetine)')

# 2d. Anxiolytics/sedatives (抗不安薬 - 不安障害 proxy)
# Note: BZD anxiolytics are in class 112
anxiolytic_kw = ['エチゾラム', 'デパス', 'アルプラゾラム', 'ソラナックス', 'コンスタン',
                 'ロラゼパム', 'ワイパックス', 'ジアゼパム', 'セルシン', 'ホリゾン',
                 'クロチアゼパム', 'リーゼ', 'ブロマゼパム', 'レキソタン',
                 'フルジアゼパム', 'エリスパン', 'メダゼパム', 'レスミット',
                 'オキサゾラム', 'セレナール', 'クロキサゾラム', 'セパゾン',
                 'タンドスピロン', 'セディール', 'ヒドロキシジン', 'アタラックス']
anxiolytic_data, anxiolytic_total = extract_drug_data_by_prefecture(
    outpatient_file, '内服薬 外来 (院外)', anxiolytic_kw, 'Anxiolytics')

# 2e. Fibromyalgia - use specific FM drugs that overlap with pregabalin
# Since pregabalin IS the main FM treatment, we can't separate easily.
# But we can note that FM prevalence varies by region and use population survey data.
# For now, we'll flag this limitation.

# ============================================================
# STEP 3: Load Phase 1 data
# ============================================================
print("\n" + "=" * 80)
print("STEP 3: Loading Phase 1 data")
print("=" * 80)

phase1_data = {}
with open(OUTPUT_DIR + 'prefecture_results.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        pc = int(row['pref_code'])
        phase1_data[pc] = {
            'pref_name': row['pref_name'],
            'region': row['region'],
            'is_tohoku': row['is_tohoku'],
            'surgery_count': float(row['surgery_count']),
            'analgesic_quantity': float(row['analgesic_quantity']),
            'population_thousands': float(row['population_thousands']),
            'analgesic_per_surgery': float(row['analgesic_per_surgery']),
        }
print(f"  Loaded Phase 1 data for {len(phase1_data)} prefectures")

# ============================================================
# STEP 4: Build integrated dataset
# ============================================================
print("\n" + "=" * 80)
print("STEP 4: Building integrated dataset")
print("=" * 80)

output_rows = []
for pc in range(1, 48):
    p1 = phase1_data[pc]
    pop = p1['population_thousands']
    surg = p1['surgery_count']
    
    row = {
        'pref_code': pc,
        'pref_name': p1['pref_name'],
        'region': p1['region'],
        'is_tohoku': p1['is_tohoku'],
        'population_thousands': pop,
        'surgery_count': surg,
        
        # Phase 1 metrics
        'analgesic_quantity_inpatient': p1['analgesic_quantity'],
        'analgesic_per_surgery': p1['analgesic_per_surgery'],
        
        # Neuropathic pain drugs (outpatient) - CPSP proxy
        'pregabalin_outpatient': pregabalin_data[pc],
        'mirogabalin_outpatient': mirogabalin_data[pc],
        'duloxetine_outpatient': duloxetine_data[pc],
        'tramadol_outpatient': tramadol_data[pc],
        'neurotropin_outpatient': neurotropin_data[pc],
        'neuropathic_total_outpatient': neuropathic_total[pc],
        
        # Confounder proxies (outpatient)
        'diabetes_drugs_outpatient': diabetes_data[pc],
        'herpes_antivirals_outpatient': herpes_data[pc],
        'antidepressants_outpatient': antidep_data[pc],
        'anxiolytics_outpatient': anxiolytic_data[pc],
        
        # Per-surgery indices
        'neuropathic_per_surgery': neuropathic_total[pc] / surg if surg > 0 else 0,
        'diabetes_per_surgery': diabetes_data[pc] / surg if surg > 0 else 0,
        'herpes_per_surgery': herpes_data[pc] / surg if surg > 0 else 0,
        'antidep_per_surgery': antidep_data[pc] / surg if surg > 0 else 0,
        'anxiolytic_per_surgery': anxiolytic_data[pc] / surg if surg > 0 else 0,
        
        # Per-capita indices (per 1000 population)
        'neuropathic_per_capita': neuropathic_total[pc] / pop if pop > 0 else 0,
        'diabetes_per_capita': diabetes_data[pc] / pop if pop > 0 else 0,
        'herpes_per_capita': herpes_data[pc] / pop if pop > 0 else 0,
        'antidep_per_capita': antidep_data[pc] / pop if pop > 0 else 0,
        'anxiolytic_per_capita': anxiolytic_data[pc] / pop if pop > 0 else 0,
    }
    output_rows.append(row)

# Save integrated dataset
outpath = OUTPUT_DIR + 'integrated_cpsp_dataset.csv'
fieldnames = list(output_rows[0].keys())
with open(outpath, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(output_rows)
print(f"  Saved integrated dataset to {outpath}")
print(f"  {len(output_rows)} prefectures x {len(fieldnames)} variables")

# Quick summary
print("\n--- Quick summary by region ---")
from collections import defaultdict
region_data = defaultdict(list)
for row in output_rows:
    region_data[row['region']].append(row)

for region in sorted(region_data.keys()):
    rows = region_data[region]
    n = len(rows)
    avg_neuro = sum(r['neuropathic_per_surgery'] for r in rows) / n
    avg_diab = sum(r['diabetes_per_surgery'] for r in rows) / n
    avg_herp = sum(r['herpes_per_surgery'] for r in rows) / n
    avg_adep = sum(r['antidep_per_surgery'] for r in rows) / n
    avg_anxi = sum(r['anxiolytic_per_surgery'] for r in rows) / n
    print(f"  {region} (n={n}): neuro/surg={avg_neuro:.1f}, diab/surg={avg_diab:.1f}, "
          f"herpes/surg={avg_herp:.1f}, antidep/surg={avg_adep:.1f}, anxio/surg={avg_anxi:.1f}")

print("\nDone extracting confounder data.")
