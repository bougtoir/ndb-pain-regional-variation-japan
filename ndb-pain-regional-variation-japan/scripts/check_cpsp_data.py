#!/usr/bin/env python3
"""Check NDB open data for chronic postsurgical pain (CPSP) proxy indicators."""
import openpyxl

print("=" * 80)
print("1. OUTPATIENT DRUGS - Searching for neuropathic pain drugs")
print("=" * 80)

wb = openpyxl.load_workbook('/home/ubuntu/analysis/data/outpatient_drugs_prefecture.xlsx', read_only=True)
for sname in wb.sheetnames:
    print(f"\nSheet: {sname}")
    sheet = wb[sname]
    targets = ['プレガバリン', 'リリカ', 'ミロガバリン', 'タリージェ', 
               'デュロキセチン', 'サインバルタ', 'ガバペンチン', 'トラマドール',
               'トラムセット', 'ノイロトロピン', 'ワクシニア']
    count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        if len(vals) > 3 and vals[3] and any(t in str(vals[3]) for t in targets):
            print(f"  Row {i}: class={vals[0]}({vals[1]}), drug={vals[3]}, total={vals[8]}")
            count += 1
    print(f"  Found {count} matching drugs in this sheet")
wb.close()

print("\n" + "=" * 80)
print("2. INPATIENT DRUGS - Searching for neuropathic pain drugs")
print("=" * 80)

wb2 = openpyxl.load_workbook('/home/ubuntu/analysis/data/inpatient_drugs_prefecture.xlsx', read_only=True)
for sname in wb2.sheetnames:
    print(f"\nSheet: {sname}")
    sheet = wb2[sname]
    count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        if len(vals) > 3 and vals[3] and any(t in str(vals[3]) for t in targets):
            print(f"  Row {i}: class={vals[0]}({vals[1]}), drug={vals[3]}, total={vals[8]}")
            count += 1
    print(f"  Found {count} matching drugs in this sheet")
wb2.close()

print("\n" + "=" * 80)
print("3. PROCEDURES (anesthesia) - Checking for nerve block data")
print("=" * 80)

wb3 = openpyxl.load_workbook('/home/ubuntu/analysis/data/anesthesia_prefecture.xlsx', read_only=True)
for sname in wb3.sheetnames:
    print(f"\nSheet: {sname}")
    sheet = wb3[sname]
    nerve_targets = ['神経ブロック', 'トリガーポイント', 'ペインクリニック', 
                     'ボトックス', 'パルス高周波', '硬膜外', 'ブロック']
    count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        for ci, v in enumerate(vals[:6]):
            if v and any(t in str(v) for t in nerve_targets):
                print(f"  Row {i}: {vals[:6]}")
                count += 1
                break
    print(f"  Found {count} matching procedures in this sheet")
wb3.close()

print("\n" + "=" * 80)
print("4. SURGERY DATA - Check for pain-related procedures")
print("=" * 80)

wb4 = openpyxl.load_workbook('/home/ubuntu/analysis/data/surgery_prefecture.xlsx', read_only=True)
for sname in wb4.sheetnames:
    print(f"\nSheet: {sname}")
    sheet = wb4[sname]
    pain_targets = ['神経ブロック', '脊髄刺激', '神経切断', '神経剥離']
    count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        for ci, v in enumerate(vals[:6]):
            if v and any(t in str(v) for t in pain_targets):
                print(f"  Row {i}: {vals[:6]}")
                count += 1
                break
    print(f"  Found {count} matching procedures in this sheet")
wb4.close()

print("\n" + "=" * 80)
print("5. OUTPATIENT DRUGS - All sheet names and drug classes (category 119)")
print("=" * 80)

wb5 = openpyxl.load_workbook('/home/ubuntu/analysis/data/outpatient_drugs_prefecture.xlsx', read_only=True)
sheet = wb5[wb5.sheetnames[0]]
class_119_count = 0
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    vals = list(row)
    if len(vals) > 0 and vals[0] == 119:
        if class_119_count < 20:
            print(f"  Row {i}: {vals[0]}({vals[1]}), drug={vals[3]}, total={vals[8]}")
        class_119_count += 1
print(f"  Total drugs in class 119 (other CNS drugs): {class_119_count}")
wb5.close()

print("\nDone.")
