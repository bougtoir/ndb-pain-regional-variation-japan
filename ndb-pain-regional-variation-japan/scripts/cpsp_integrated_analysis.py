#!/usr/bin/env python3
"""
CPSP Integrated Analysis: Phase 2
Neuropathic pain drug prescriptions (outpatient) as CPSP proxy,
adjusted for confounding diseases, integrated with Phase 1 acute pain data.
"""
import openpyxl
import csv
import os
import json
import numpy as np
from collections import defaultdict

OUTPUT_DIR = '/home/ubuntu/analysis/output/'
DATA_DIR = '/home/ubuntu/analysis/data/'

PREF_NAMES = {
    1:'北海道',2:'青森県',3:'岩手県',4:'宮城県',5:'秋田県',6:'山形県',7:'福島県',
    8:'茨城県',9:'栃木県',10:'群馬県',11:'埼玉県',12:'千葉県',13:'東京都',14:'神奈川県',
    15:'新潟県',16:'富山県',17:'石川県',18:'福井県',19:'山梨県',20:'長野県',
    21:'岐阜県',22:'静岡県',23:'愛知県',24:'三重県',25:'滋賀県',26:'京都府',
    27:'大阪府',28:'兵庫県',29:'奈良県',30:'和歌山県',31:'鳥取県',32:'島根県',
    33:'岡山県',34:'広島県',35:'山口県',36:'徳島県',37:'香川県',38:'愛媛県',
    39:'高知県',40:'福岡県',41:'佐賀県',42:'長崎県',43:'熊本県',44:'大分県',
    45:'宮崎県',46:'鹿児島県',47:'沖縄県'
}

REGION_MAP = {
    1:'北海道', 2:'東北',3:'東北',4:'東北',5:'東北',6:'東北',7:'東北',
    8:'関東',9:'関東',10:'関東',11:'関東',12:'関東',13:'関東',14:'関東',
    15:'北陸・甲信越',16:'北陸・甲信越',17:'北陸・甲信越',18:'北陸・甲信越',19:'北陸・甲信越',20:'北陸・甲信越',
    21:'東海',22:'東海',23:'東海',24:'東海',
    25:'近畿',26:'近畿',27:'近畿',28:'近畿',29:'近畿',30:'近畿',
    31:'中国',32:'中国',33:'中国',34:'中国',35:'中国',
    36:'四国',37:'四国',38:'四国',39:'四国',
    40:'九州・沖縄',41:'九州・沖縄',42:'九州・沖縄',43:'九州・沖縄',44:'九州・沖縄',45:'九州・沖縄',46:'九州・沖縄',47:'九州・沖縄'
}

PREF_COL_START = 6  # 0-indexed column for 北海道 in anesthesia file (col 6)
DRUG_PREF_COL_START = 9  # 0-indexed column for 北海道 in drug file (col 9)

def safe_float(val):
    if val is None or val == '-' or val == '－' or val == '‐':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

# ============================================================
# EXTRACTION FUNCTIONS
# ============================================================

def extract_drugs_by_keywords(filepath, sheet_name, keywords, label):
    """Extract drug prescription quantities by prefecture using drug name keywords."""
    print(f"\n  Extracting: {label}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb[sheet_name]
    
    pref_totals = {i: 0.0 for i in range(1, 48)}
    national = 0.0
    count = 0
    
    for row in sheet.iter_rows(values_only=True):
        vals = list(row)
        if len(vals) < 56:
            continue
        drug_name = str(vals[3]) if vals[3] else ''
        if any(kw in drug_name for kw in keywords):
            count += 1
            national += safe_float(vals[8])
            for pc in range(1, 48):
                col = DRUG_PREF_COL_START + (pc - 1)
                if col < len(vals):
                    pref_totals[pc] += safe_float(vals[col])
    wb.close()
    print(f"    -> {count} entries, national total: {national:,.0f}")
    return pref_totals, national, count

def extract_nerve_blocks_by_prefecture(filepath, sheet_name):
    """Extract outpatient nerve block procedures by prefecture."""
    print(f"\n  Extracting nerve blocks from: {sheet_name}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb[sheet_name]
    
    # Nerve block keywords
    block_kw = ['神経ブロック', '硬膜外ブロック', 'トリガーポイント', '星状神経節',
                '肩甲上神経', '腕神経叢', '肋間神経', '坐骨神経', '大腿神経',
                '閉鎖神経', '後頭神経', '眼窩上神経', '眼窩下神経',
                'ブロック（その他）', 'ブロック(その他)']
    
    pref_totals = {i: 0.0 for i in range(1, 48)}
    national = 0.0
    count = 0
    
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        vals = list(row)
        if i < 4 or len(vals) < 53:
            continue
        proc_name = str(vals[3]) if vals[3] else ''
        if any(kw in proc_name for kw in block_kw):
            count += 1
            national += safe_float(vals[5])
            for pc in range(1, 48):
                col = PREF_COL_START + (pc - 1)
                if col < len(vals):
                    pref_totals[pc] += safe_float(vals[col])
    wb.close()
    print(f"    -> {count} procedure entries, national total: {national:,.0f}")
    return pref_totals, national, count

# ============================================================
# MAIN EXTRACTION
# ============================================================
print("=" * 80)
print("CPSP INTEGRATED ANALYSIS - DATA EXTRACTION")
print("=" * 80)

outpatient_file = DATA_DIR + 'outpatient_drugs_prefecture.xlsx'
anesthesia_file = DATA_DIR + 'anesthesia_prefecture.xlsx'
sheet_name = '内服薬 外来 (院外)'

# --- CPSP proxy drugs (outpatient) ---
print("\n[1] CPSP proxy: Neuropathic pain drugs (outpatient)")

pregabalin_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['プレガバリン', 'リリカ'], 'Pregabalin')

mirogabalin_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['ミロガバリン', 'タリージェ'], 'Mirogabalin')

duloxetine_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['デュロキセチン', 'サインバルタ'], 'Duloxetine')

tramadol_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['トラマドール', 'トラムセット'], 'Tramadol')

neurotropin_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['ノイロトロピン', 'ワクシニア'], 'Neurotropin')

# --- Confounder proxies ---
print("\n[2] Confounder proxies")

# 2a. Diabetes drugs (all oral hypoglycemics)
diabetes_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['メトホルミン', 'グリメピリド', 'アマリール', 'シタグリプチン', 'ジャヌビア',
     'リナグリプチン', 'トラゼンタ', 'テネリグリプチン', 'テネリア', 'アログリプチン',
     'ネシーナ', 'ビルダグリプチン', 'エクア', 'サキサグリプチン', 'オングリザ',
     'エンパグリフロジン', 'ジャディアンス', 'ダパグリフロジン', 'フォシーガ',
     'カナグリフロジン', 'カナグル', 'イプラグリフロジン', 'スーグラ', 'ルセオグリフロジン',
     'ピオグリタゾン', 'アクトス', 'ボグリボース', 'ベイスン', 'ミグリトール', 'セイブル',
     'グリクラジド', 'グリベンクラミド', 'ダオニール', 'オイグルコン',
     'レパグリニド', 'シュアポスト', 'ナテグリニド', 'スターシス', 'ファスティック',
     'ミチグリニド', 'グルファスト', 'メトグルコ', 'グルベス',
     'マリゼブ', 'オマリグリプチン', 'ザファテック', 'トレラグリプチン'],
    'Diabetes drugs (oral hypoglycemics)')

# 2b. Herpes zoster antivirals
herpes_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['バラシクロビル', 'バルトレックス', 'アシクロビル', 'ゾビラックス',
     'ファムシクロビル', 'ファムビル', 'アメナメビル', 'アメナリーフ'],
    'Herpes zoster antivirals')

# 2c. Antidepressants (excluding duloxetine - already in CPSP proxy)
antidep_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['パロキセチン', 'パキシル', 'セルトラリン', 'ジェイゾロフト',
     'エスシタロプラム', 'レクサプロ', 'フルボキサミン', 'デプロメール', 'ルボックス',
     'ミルナシプラン', 'トレドミン', 'ベンラファキシン', 'イフェクサー',
     'ミルタザピン', 'リフレックス', 'レメロン',
     'アミトリプチリン', 'トリプタノール', 'ノリトレン', 'イミプラミン', 'トフラニール',
     'クロミプラミン', 'アナフラニール', 'ボルチオキセチン', 'トリンテリックス'],
    'Antidepressants (excl. duloxetine)')

# 2d. Anxiolytics
anxiolytic_data, _, _ = extract_drugs_by_keywords(
    outpatient_file, sheet_name,
    ['エチゾラム', 'デパス', 'アルプラゾラム', 'ソラナックス', 'コンスタン',
     'ロラゼパム', 'ワイパックス', 'ジアゼパム', 'セルシン', 'ホリゾン',
     'クロチアゼパム', 'リーゼ', 'ブロマゼパム', 'レキソタン',
     'フルジアゼパム', 'エリスパン', 'メダゼパム', 'レスミット',
     'オキサゾラム', 'セレナール', 'クロキサゾラム', 'セパゾン',
     'タンドスピロン', 'セディール', 'ヒドロキシジン', 'アタラックス'],
    'Anxiolytics')

# --- Nerve blocks (outpatient) ---
print("\n[3] Outpatient nerve blocks")
nerve_block_data, _, _ = extract_nerve_blocks_by_prefecture(anesthesia_file, '外来')

# ============================================================
# LOAD PHASE 1 DATA
# ============================================================
print("\n[4] Loading Phase 1 data")
phase1 = {}
with open(OUTPUT_DIR + 'prefecture_results.csv', 'r', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        pc = int(row['pref_code'])
        phase1[pc] = {
            'surgery_count': float(row['surgery_count']),
            'analgesic_quantity': float(row['analgesic_quantity']),
            'population_thousands': float(row['population_thousands']),
            'analgesic_per_surgery': float(row['analgesic_per_surgery']),
        }
print(f"  Loaded {len(phase1)} prefectures")

# ============================================================
# BUILD INTEGRATED DATASET
# ============================================================
print("\n[5] Building integrated dataset")

rows = []
for pc in range(1, 48):
    p1 = phase1[pc]
    pop = p1['population_thousands']
    surg = p1['surgery_count']
    
    # Combine neuropathic pain drugs
    neuro_total = (pregabalin_data[pc] + mirogabalin_data[pc] + 
                   duloxetine_data[pc] + tramadol_data[pc] + neurotropin_data[pc])
    
    # Core neuropathic only (pregabalin + mirogabalin) - least confounded
    core_neuro = pregabalin_data[pc] + mirogabalin_data[pc]
    
    row = {
        'pref_code': pc,
        'pref_name': PREF_NAMES[pc],
        'region': REGION_MAP[pc],
        'is_tohoku': 1 if 2 <= pc <= 7 else 0,
        'population_thousands': pop,
        'surgery_count': surg,
        
        # Phase 1: Acute pain proxy
        'acute_analgesic_per_surgery': p1['analgesic_per_surgery'],
        
        # CPSP proxy: Neuropathic pain drugs (outpatient)
        'pregabalin_out': pregabalin_data[pc],
        'mirogabalin_out': mirogabalin_data[pc],
        'duloxetine_out': duloxetine_data[pc],
        'tramadol_out': tramadol_data[pc],
        'neurotropin_out': neurotropin_data[pc],
        'neuropathic_total_out': neuro_total,
        'core_neuropathic_out': core_neuro,
        
        # Nerve blocks (outpatient)
        'nerve_block_out': nerve_block_data[pc],
        
        # Confounders
        'diabetes_drugs_out': diabetes_data[pc],
        'herpes_antivirals_out': herpes_data[pc],
        'antidepressants_out': antidep_data[pc],
        'anxiolytics_out': anxiolytic_data[pc],
        
        # Per-surgery indices
        'neuropathic_per_surgery': neuro_total / surg if surg > 0 else 0,
        'core_neuro_per_surgery': core_neuro / surg if surg > 0 else 0,
        'nerve_block_per_surgery': nerve_block_data[pc] / surg if surg > 0 else 0,
        'diabetes_per_surgery': diabetes_data[pc] / surg if surg > 0 else 0,
        'herpes_per_surgery': herpes_data[pc] / surg if surg > 0 else 0,
        'antidep_per_surgery': antidep_data[pc] / surg if surg > 0 else 0,
        'anxiolytic_per_surgery': anxiolytic_data[pc] / surg if surg > 0 else 0,
        
        # Per-capita indices
        'neuropathic_per_capita': neuro_total / pop if pop > 0 else 0,
        'diabetes_per_capita': diabetes_data[pc] / pop if pop > 0 else 0,
        'herpes_per_capita': herpes_data[pc] / pop if pop > 0 else 0,
    }
    rows.append(row)

# Save
outpath = OUTPUT_DIR + 'cpsp_integrated_dataset.csv'
with open(outpath, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
print(f"  Saved: {outpath} ({len(rows)} rows x {len(rows[0])} cols)")

# ============================================================
# REGRESSION ANALYSIS
# ============================================================
print("\n" + "=" * 80)
print("REGRESSION ANALYSIS")
print("=" * 80)

# Convert to numpy arrays
n = len(rows)
Y_neuro = np.array([r['neuropathic_per_surgery'] for r in rows])
Y_core = np.array([r['core_neuro_per_surgery'] for r in rows])
Y_block = np.array([r['nerve_block_per_surgery'] for r in rows])
X_diabetes = np.array([r['diabetes_per_surgery'] for r in rows])
X_herpes = np.array([r['herpes_per_surgery'] for r in rows])
X_antidep = np.array([r['antidep_per_surgery'] for r in rows])
X_anxio = np.array([r['anxiolytic_per_surgery'] for r in rows])
X_acute = np.array([r['acute_analgesic_per_surgery'] for r in rows])
is_tohoku = np.array([r['is_tohoku'] for r in rows])
regions = [r['region'] for r in rows]
pref_names = [r['pref_name'] for r in rows]

# Simple OLS regression using numpy
def ols_regression(Y, X_matrix, var_names):
    """Run OLS regression: Y = X @ beta + epsilon"""
    n = len(Y)
    k = X_matrix.shape[1]
    
    # Add intercept
    X = np.column_stack([np.ones(n), X_matrix])
    
    # OLS: beta = (X'X)^-1 X'Y
    XtX_inv = np.linalg.inv(X.T @ X)
    beta = XtX_inv @ X.T @ Y
    
    # Residuals and statistics
    Y_hat = X @ beta
    residuals = Y - Y_hat
    SSE = residuals @ residuals
    SST = np.sum((Y - np.mean(Y))**2)
    R2 = 1 - SSE / SST
    R2_adj = 1 - (1 - R2) * (n - 1) / (n - k - 1)
    
    # Standard errors
    sigma2 = SSE / (n - k - 1)
    se = np.sqrt(np.diag(sigma2 * XtX_inv))
    t_stats = beta / se
    
    # P-values (two-tailed, using normal approximation for large n)
    from scipy import stats as sp_stats
    p_values = 2 * (1 - sp_stats.t.cdf(np.abs(t_stats), df=n-k-1))
    
    names = ['Intercept'] + var_names
    print(f"\n  R² = {R2:.4f}, Adjusted R² = {R2_adj:.4f}")
    print(f"  {'Variable':<25} {'Coef':>10} {'SE':>10} {'t':>8} {'p':>8}")
    print(f"  {'-'*65}")
    for i, name in enumerate(names):
        sig = ''
        if p_values[i] < 0.001: sig = '***'
        elif p_values[i] < 0.01: sig = '**'
        elif p_values[i] < 0.05: sig = '*'
        print(f"  {name:<25} {beta[i]:>10.4f} {se[i]:>10.4f} {t_stats[i]:>8.3f} {p_values[i]:>8.4f} {sig}")
    
    return beta, se, t_stats, p_values, R2, R2_adj, residuals, Y_hat

# --- Model 1: Unadjusted neuropathic pain drugs per surgery ---
print("\n--- Model 1: Neuropathic pain / surgery (unadjusted) ---")
print(f"  Mean = {np.mean(Y_neuro):.1f}, SD = {np.std(Y_neuro):.1f}")
print(f"  Tohoku mean = {np.mean(Y_neuro[is_tohoku==1]):.1f}")
print(f"  Non-Tohoku mean = {np.mean(Y_neuro[is_tohoku==0]):.1f}")

# T-test Tohoku vs others
from scipy import stats
t_unadj, p_unadj = stats.ttest_ind(Y_neuro[is_tohoku==1], Y_neuro[is_tohoku==0])
# Pooled within-group SD for Cohen's d
_tohoku_vals = Y_neuro[is_tohoku==1]
_rest_vals = Y_neuro[is_tohoku==0]
_pooled_std = np.sqrt(((len(_tohoku_vals)-1)*np.var(_tohoku_vals, ddof=1) + (len(_rest_vals)-1)*np.var(_rest_vals, ddof=1)) / (len(_tohoku_vals)+len(_rest_vals)-2))
d_unadj = (np.mean(_tohoku_vals) - np.mean(_rest_vals)) / _pooled_std
print(f"  T-test: t={t_unadj:.3f}, p={p_unadj:.4f}, Cohen's d={d_unadj:.3f}")

# --- Model 2: Adjusted for confounders ---
print("\n--- Model 2: Neuropathic pain / surgery ~ confounders + is_tohoku ---")
X_confounders = np.column_stack([X_diabetes, X_herpes, X_antidep, X_anxio, is_tohoku])
conf_names = ['Diabetes/surg', 'Herpes/surg', 'Antidep/surg', 'Anxiolytic/surg', 'is_Tohoku']
m2_results = ols_regression(Y_neuro, X_confounders, conf_names)
m2_residuals = m2_results[6]

# --- Model 3: Core neuropathic (pregabalin + mirogabalin only) adjusted ---
print("\n--- Model 3: Core neuropathic (PGB+MGB) / surgery ~ confounders + is_tohoku ---")
X_conf3 = np.column_stack([X_diabetes, X_herpes, X_antidep, X_anxio, is_tohoku])
m3_results = ols_regression(Y_core, X_conf3, conf_names)

# --- Model 4: Nerve blocks adjusted ---
print("\n--- Model 4: Nerve blocks / surgery ~ confounders + is_tohoku ---")
m4_results = ols_regression(Y_block, np.column_stack([X_diabetes, X_herpes, X_antidep, X_anxio, is_tohoku]),
                            conf_names)

# --- Model 5: Integration with Phase 1 acute pain ---
print("\n--- Model 5: Neuropathic pain ~ Acute pain + confounders ---")
X_int = np.column_stack([X_acute, X_diabetes, X_herpes, X_antidep, X_anxio, is_tohoku])
int_names = ['Acute_pain/surg', 'Diabetes/surg', 'Herpes/surg', 'Antidep/surg', 'Anxiolytic/surg', 'is_Tohoku']
m5_results = ols_regression(Y_neuro, X_int, int_names)

# ============================================================
# ADJUSTED CPSP INDEX (residuals after removing confounder effects)
# ============================================================
print("\n" + "=" * 80)
print("ADJUSTED CPSP INDEX (confounder-removed residuals)")
print("=" * 80)

# Regress neuropathic on confounders only (without is_tohoku)
X_conf_only = np.column_stack([X_diabetes, X_herpes, X_antidep, X_anxio])
conf_only_names = ['Diabetes/surg', 'Herpes/surg', 'Antidep/surg', 'Anxiolytic/surg']
print("\n--- Confounder-only model ---")
adj_results = ols_regression(Y_neuro, X_conf_only, conf_only_names)
adj_residuals = adj_results[6]
adj_fitted = adj_results[7]

# The residuals = "unexplained" neuropathic pain after accounting for diabetes, herpes, depression, anxiety
# This is our adjusted CPSP index
adjusted_cpsp = adj_residuals  # Positive = more neuropathic pain than expected given confounders

# Also do the same for core neuropathic
adj_core_results = ols_regression(Y_core, X_conf_only, conf_only_names)
adj_core_residuals = adj_core_results[6]

print("\n--- Adjusted CPSP Index by Region ---")
region_adj = defaultdict(list)
for i, row in enumerate(rows):
    region_adj[row['region']].append(adjusted_cpsp[i])

for region in sorted(region_adj.keys(), key=lambda x: np.mean(region_adj[x])):
    vals = region_adj[region]
    print(f"  {region}: mean={np.mean(vals):+.1f}, SD={np.std(vals):.1f}, n={len(vals)}")

# Tohoku adjusted test
print(f"\n  Tohoku adjusted mean: {np.mean(adjusted_cpsp[is_tohoku==1]):+.1f}")
print(f"  Non-Tohoku adjusted mean: {np.mean(adjusted_cpsp[is_tohoku==0]):+.1f}")
t_adj, p_adj = stats.ttest_ind(adjusted_cpsp[is_tohoku==1], adjusted_cpsp[is_tohoku==0])
# Pooled within-group SD for Cohen's d
_tohoku_adj = adjusted_cpsp[is_tohoku==1]
_rest_adj = adjusted_cpsp[is_tohoku==0]
_pooled_std_adj = np.sqrt(((len(_tohoku_adj)-1)*np.var(_tohoku_adj, ddof=1) + (len(_rest_adj)-1)*np.var(_rest_adj, ddof=1)) / (len(_tohoku_adj)+len(_rest_adj)-2))
d_adj = (np.mean(_tohoku_adj) - np.mean(_rest_adj)) / _pooled_std_adj
print(f"  T-test (adjusted): t={t_adj:.3f}, p={p_adj:.4f}, Cohen's d={d_adj:.3f}")

# ============================================================
# SAVE FULL RESULTS
# ============================================================
print("\n[6] Saving full results")

# Add adjusted CPSP index to rows
for i, row in enumerate(rows):
    row['adjusted_cpsp_index'] = adjusted_cpsp[i]
    row['adjusted_core_cpsp_index'] = adj_core_residuals[i]
    row['predicted_neuropathic'] = adj_fitted[i]

outpath2 = OUTPUT_DIR + 'cpsp_integrated_results.csv'
with open(outpath2, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
print(f"  Saved: {outpath2}")

# Save regression summary
summary = {
    'model1_unadjusted': {
        'tohoku_mean': float(np.mean(Y_neuro[is_tohoku==1])),
        'non_tohoku_mean': float(np.mean(Y_neuro[is_tohoku==0])),
        't_statistic': float(t_unadj),
        'p_value': float(p_unadj),
        'cohens_d': float(d_unadj),
    },
    'model2_adjusted': {
        'R2': float(m2_results[4]),
        'R2_adj': float(m2_results[5]),
        'tohoku_coef': float(m2_results[0][-1]),
        'tohoku_p': float(m2_results[3][-1]),
    },
    'model3_core_neuropathic': {
        'R2': float(m3_results[4]),
        'R2_adj': float(m3_results[5]),
        'tohoku_coef': float(m3_results[0][-1]),
        'tohoku_p': float(m3_results[3][-1]),
    },
    'model4_nerve_blocks': {
        'R2': float(m4_results[4]),
        'R2_adj': float(m4_results[5]),
        'tohoku_coef': float(m4_results[0][-1]),
        'tohoku_p': float(m4_results[3][-1]),
    },
    'model5_integrated': {
        'R2': float(m5_results[4]),
        'R2_adj': float(m5_results[5]),
        'acute_pain_coef': float(m5_results[0][1]),
        'acute_pain_p': float(m5_results[3][1]),
        'tohoku_coef': float(m5_results[0][-1]),
        'tohoku_p': float(m5_results[3][-1]),
    },
    'adjusted_cpsp_test': {
        'tohoku_mean': float(np.mean(adjusted_cpsp[is_tohoku==1])),
        'non_tohoku_mean': float(np.mean(adjusted_cpsp[is_tohoku==0])),
        't_statistic': float(t_adj),
        'p_value': float(p_adj),
        'cohens_d': float(d_adj),
    }
}

with open(OUTPUT_DIR + 'cpsp_regression_summary.json', 'w') as f:
    json.dump(summary, f, indent=2)
print(f"  Saved regression summary")

# ============================================================
# CORRELATION MATRIX
# ============================================================
print("\n" + "=" * 80)
print("CORRELATION MATRIX")
print("=" * 80)

var_data = {
    'Acute pain/surg': X_acute,
    'Neuropathic/surg': Y_neuro,
    'Core neuro/surg': Y_core,
    'Nerve block/surg': Y_block,
    'Diabetes/surg': X_diabetes,
    'Herpes/surg': X_herpes,
    'Antidep/surg': X_antidep,
    'Anxiolytic/surg': X_anxio,
    'Adj CPSP index': adjusted_cpsp,
}

var_names_list = list(var_data.keys())
print(f"\n  {'':>20}", end='')
for vn in var_names_list:
    print(f" {vn[:12]:>12}", end='')
print()

for vn1 in var_names_list:
    print(f"  {vn1:>20}", end='')
    for vn2 in var_names_list:
        r, p = stats.pearsonr(var_data[vn1], var_data[vn2])
        sig = '*' if p < 0.05 else ' '
        print(f" {r:>11.3f}{sig}", end='')
    print()

# Phase 1 vs Phase 2 correlation
r_acute_chronic, p_ac = stats.pearsonr(X_acute, Y_neuro)
r_acute_adj, p_aa = stats.pearsonr(X_acute, adjusted_cpsp)
print(f"\n  Phase 1 (acute) vs Phase 2 (chronic) correlation:")
print(f"    Raw: r={r_acute_chronic:.3f}, p={p_ac:.4f}")
print(f"    Adjusted: r={r_acute_adj:.3f}, p={p_aa:.4f}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
